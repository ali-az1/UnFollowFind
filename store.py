"""Persistence: whitelist, xlsx exports and run-to-run snapshots.

Snapshots replace the old prefollowers.xlsx trick: every run drops a timestamped
JSON file under data/snapshots/, and the next run diffs against it automatically.
Everything is keyed on the numeric user id, so a username change is reported as a
rename instead of a fake "lost follower + new follower" pair.
"""

from __future__ import annotations

import json
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable

DATA_DIR = Path("data")
SNAPSHOT_DIR = DATA_DIR / "snapshots"
WHITELIST_FILE = Path("whitelist.txt")

User = dict[str, Any]


# ---------------------------------------------------------------- whitelist

def load_whitelist(path: str | Path = WHITELIST_FILE) -> set[str]:
    """Usernames to ignore in the 'not following you back' report (one per line)."""
    path = Path(path)
    if not path.exists():
        return set()
    names = set()
    for line in path.read_text(encoding="utf-8").splitlines():
        line = line.split("#", 1)[0].strip().lstrip("@")
        if line:
            names.add(line.lower())
    return names


# -------------------------------------------------------------------- xlsx

def load_xlsx(path: str | Path) -> list[User]:
    """Read a username list back out of an .xlsx.

    Handles both this tool's exports and the older username-only sheets, which have
    a single column and therefore no account ids.
    """
    import openpyxl

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)

    header = next(rows, None) or ()
    columns = {
        str(cell).strip().lower().replace(" ", "_"): index
        for index, cell in enumerate(header) if cell
    }
    if "username" not in columns:  # no header row -- first column is the username
        columns = {"username": 0}
        rows = sheet.iter_rows(values_only=True)

    users: list[User] = []
    for row in rows:
        def value(name: str, default: Any = "") -> Any:
            index = columns.get(name)
            if index is None or index >= len(row) or row[index] is None:
                return default
            return row[index]

        username = str(value("username")).strip().lstrip("@")
        if not username:
            continue
        users.append({
            "pk": str(value("pk")).strip(),
            "username": username,
            "full_name": str(value("full_name")),
            "is_private": bool(value("is_private", False)),
            "is_verified": bool(value("is_verified", False)),
        })
    workbook.close()
    return users


def load_baseline(paths: Iterable[str | Path]) -> dict[str, Any]:
    """Build a comparison baseline from snapshot .json and/or exported .xlsx files.

    A file's role is taken from its name/sheet: anything mentioning "following" is a
    following list, everything else is a follower list (so prefollowers.xlsx works).
    """
    baseline: dict[str, Any] = {"followers": None, "following": None, "sources": []}

    for path in paths:
        path = Path(path)
        if not path.exists():
            raise FileNotFoundError(f"Baseline file not found: {path}")
        stamp = datetime.fromtimestamp(path.stat().st_mtime).isoformat(timespec="seconds")

        if path.suffix.lower() == ".json":
            snapshot = json.loads(path.read_text(encoding="utf-8"))
            baseline["followers"] = snapshot.get("followers")
            baseline["following"] = snapshot.get("following")
            baseline["sources"].append(f"{path.name} ({snapshot.get('taken_at', stamp)})")
            continue

        kind = "following" if "following" in path.stem.lower() else "followers"
        baseline[kind] = load_xlsx(path)
        baseline["sources"].append(f"{path.name} -> {kind}, {stamp}")

    return baseline


def save_xlsx(path: str | Path, users: Iterable[User], sheet: str = "Sheet1") -> Path:
    import openpyxl

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = openpyxl.Workbook()
    sheet_obj = workbook.active
    sheet_obj.title = sheet[:31]
    columns = ["username", "full_name", "pk", "is_private", "is_verified"]
    sheet_obj.append([c.replace("_", " ").title() for c in columns])
    for user in users:
        sheet_obj.append([user.get(c, "") for c in columns])
    sheet_obj.freeze_panes = "A2"
    for index, width in enumerate((28, 32, 16, 12, 12), start=1):
        sheet_obj.column_dimensions[chr(64 + index)].width = width
    workbook.save(path)
    return path


# --------------------------------------------------------------- snapshots

def save_snapshot(
    username: str,
    followers: list[User],
    following: list[User],
    directory: str | Path = SNAPSHOT_DIR,
) -> Path:
    directory = Path(directory)
    directory.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now(timezone.utc).strftime("%Y%m%d-%H%M%S")
    path = directory / f"{username}-{stamp}.json"
    path.write_text(
        json.dumps(
            {
                "username": username,
                "taken_at": datetime.now(timezone.utc).isoformat(timespec="seconds"),
                "followers": followers,
                "following": following,
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    return path


def latest_snapshot(
    username: str, directory: str | Path = SNAPSHOT_DIR, exclude: Path | None = None
) -> dict[str, Any] | None:
    directory = Path(directory)
    if not directory.exists():
        return None
    candidates = sorted(
        (p for p in directory.glob(f"{username}-*.json") if p != exclude),
        key=lambda p: p.name,
    )
    if not candidates:
        return None
    snapshot = json.loads(candidates[-1].read_text(encoding="utf-8"))
    snapshot["path"] = candidates[-1]
    return snapshot


def prune_snapshots(username: str, keep: int, directory: str | Path = SNAPSHOT_DIR) -> None:
    directory = Path(directory)
    if keep <= 0 or not directory.exists():
        return
    files = sorted(directory.glob(f"{username}-*.json"), key=lambda p: p.name)
    for stale in files[:-keep]:
        stale.unlink()


# ----------------------------------------------------------------- diffing

def by_pk(users: Iterable[User]) -> dict[str, User] | None:
    """Key a list by account id, or None if any entry is missing one."""
    keyed: dict[str, User] = {}
    for user in users:
        pk = str(user.get("pk") or "")
        if not pk:
            return None
        keyed[pk] = user
    return keyed


def by_username(users: Iterable[User]) -> dict[str, User]:
    return {u["username"].lower(): u for u in users if u.get("username")}


def diff(previous: Iterable[User], current: Iterable[User]) -> dict[str, list[User]]:
    """Compare two user lists: what appeared, what disappeared, who renamed.

    Matching is done on account ids when both sides have them, which is what makes a
    username change read as a rename instead of a departure plus an arrival. Older
    username-only exports fall back to matching on the handle.
    """
    previous, current = list(previous), list(current)
    old, new = by_pk(previous), by_pk(current)
    by_id = old is not None and new is not None
    if not by_id:
        old, new = by_username(previous), by_username(current)

    added = [new[pk] for pk in new.keys() - old.keys()]
    removed = [old[pk] for pk in old.keys() - new.keys()]
    renamed = [
        {**new[key], "was": old[key].get("username", "")}
        for key in (old.keys() & new.keys() if by_id else ())
        if old[key].get("username") and old[key]["username"] != new[key].get("username")
    ]
    for group in (added, removed, renamed):
        group.sort(key=lambda u: (u.get("username") or "").lower())
    return {"added": added, "removed": removed, "renamed": renamed, "matched_by_id": by_id}
