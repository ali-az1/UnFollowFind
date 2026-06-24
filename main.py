import openpyxl
from playwright.sync_api import sync_playwright
from  time import sleep
import os
AUTH = "auth.json"
import re
from bs4 import BeautifulSoup

def loader(web):
    page.locator("a").filter(has_text=f"{web}").first.click()
    RESERVED = {"explore", "reels", "reel", "p", "stories", "direct",
                "accounts", "about", "tv", "developer",
                "popular", "legal", "privacy", "terms",
                "emails", "session", "lite", "web", "your_activity"}
    dialog = page.locator('div[role="dialog"]')
    dialog.wait_for(state="visible")
    dialog.locator('a[href^="/"]').first.wait_for(state="visible")

    seen = set()
    username = []
    stable = 0
    prev = 0
    while stable < 3:

        html = dialog.inner_html()
        soup = BeautifulSoup(html, "html.parser")

        for a in soup.find_all("a", href=True):
            m = re.fullmatch(r"/([A-Za-z0-9._]+)/", a["href"])
            if m and m.group(1) not in RESERVED:
                u = m.group(1)
                if u not in seen:
                    seen.add(u)
                    username.append(u)
        #print(username)
        rows = dialog.locator('a[href^="/"]')
        if rows.count():
            rows.nth(rows.count() - 1).scroll_into_view_if_needed()
        sleep(2)                         # ← let the next batch load

        if len(seen) == prev:
            stable += 1
        else:
            stable = 0
        prev = len(seen)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{web}"
    ws["A1"] = "Username"
    for i, value in enumerate(username, start=2):
        ws[f"A{i}"] = value
    wb.save(f"{web}.xlsx")
    print(f"Saved {len(username)} usernames to {web}.xlsx")

def log_in(context):
    page = context.new_page()
    page.goto("https://www.instagram.com/")
    username=input("enter the username")
    page.get_by_role("textbox", name="email").fill(username)
    password=input("enter the password")
    page.get_by_role("textbox", name="password").fill(password)
    page.get_by_text("Log in", exact=True).click()
    sleep(5)
    if page.locator("#_r_3_").count()>0 and page.locator("#_r_3_").is_visible():
        a=input("enter the verification code")
        page.locator("#_r_3_").fill(a)
    page.get_by_text("Continue", exact=True).click()
    page.get_by_role("button",name="Save info").click()
    sleep(2)
    context.storage_state(path=AUTH)

def read_usernames(filename):
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    names = []
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        if row[0]:
            names.append(row[0])
    return names

def differ(followers,following):
    list1=[]
    for j in following:
        if j not in followers:
            list1.append(j)
    return list1






with sync_playwright() as playwright:
    browser = playwright.chromium.launch(headless=False)
    if os.path.exists(AUTH):
        context = browser.new_context(storage_state=AUTH)
    else:
        context = browser.new_context()
        log_in(context)
    page = context.new_page()
    page.goto("https://www.instagram.com/")
    sleep(2)
    if page.get_by_role("button",name="Continue").count()>0:
        page.get_by_role("button",name="Continue").first.click()
        sleep(2)
        if page.locator('input[name="pass"]').count()>0:
            page.locator('input[name="pass"]').fill(input("enter the verification code"))
            if page.get_by_text("Log in",exact=True).count()>0:
                page.get_by_text("Log in",exact=True).click()
                sleep(3)
                code_box = page.get_by_role("textbox", name=re.compile("code|security", re.I))
                if code_box.count() > 0:
                    code_box.first.fill(input("enter verification code: "))
                    page.get_by_text("Continue",exact=True).click()

    page.get_by_role("button", name="Not Now").click()
    page.goto("https://www.instagram.com/ali_zdi1/")
    loader("following")
    page.keyboard.press("Escape")
    sleep(2)
    loader("followers")
followers = read_usernames("followers.xlsx")
following = read_usernames("following.xlsx")
notfollow=differ(followers,following)
print(notfollow)

